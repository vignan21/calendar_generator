import os
import re
import glob
import hashlib
from datetime import datetime, time, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit as st


# ==========================================================
# CONFIG
# ==========================================================
INPUT_GLOB = "*.xlsx"   # all student timetable files
OUTPUT_XLSX = "student_matrix_calendar.xlsx"

# Grid resolution (use 15 if you want finer slots)
TIME_STEP_MIN = 15

DAY_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
SHEET_MATRIX = "Matrix"
SHEET_LEGEND = "Legend"


# ==========================================================
# HELPERS
# ==========================================================
DAY_ALIASES = {
    "mon": "Monday", "monday": "Monday",
    "tue": "Tuesday", "tues": "Tuesday", "tuesday": "Tuesday",
    "wed": "Wednesday", "weds": "Wednesday", "wednesday": "Wednesday",
    "thu": "Thursday", "thur": "Thursday", "thurs": "Thursday", "thursday": "Thursday",
    "fri": "Friday", "friday": "Friday",
    "sat": "Saturday", "saturday": "Saturday",
    "sun": "Sunday", "sunday": "Sunday",
}

def normalize_day(val) -> str:
    s = str(val).strip().lower()
    s = re.sub(r"[^a-z]", "", s)
    if s in DAY_ALIASES:
        return DAY_ALIASES[s]
    if len(s) >= 3 and s[:3] in DAY_ALIASES:
        return DAY_ALIASES[s[:3]]
    raise ValueError(f"Unrecognized day: {val!r}")

def parse_time(val) -> time:
    """Accept Excel time/datetime or strings like '08:30', '8.30', '08:30:00'."""
    if pd.isna(val):
        return None
    if isinstance(val, datetime):
        return val.time().replace(second=0)
    if isinstance(val, time):
        return val.replace(second=0)

    s = str(val).strip().replace(".", ":")
    m = re.match(r"^\s*(\d{1,2}):(\d{2})(?::(\d{2}))?\s*$", s)
    if not m:
        raise ValueError(f"Unrecognized time format: {val!r}")
    hh = int(m.group(1))
    mm = int(m.group(2))
    return time(hh, mm, 0)

def standardize_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Expected: 4 columns (course, day, start, end) OR headers containing those meanings.
    Fallback: uses first 4 columns.
    """
    if df.shape[1] < 4:
        raise ValueError("Input must have at least 4 columns: course, day, start, end")

    # if unnamed columns / weird headers -> first 4
    if df.columns.to_list()[:4] == list(range(4)) or all(str(c).startswith("Unnamed") for c in df.columns):
        out = df.iloc[:, :4].copy()
        out.columns = ["course", "day", "start", "end"]
        return out

    cols = {str(c).strip().lower(): c for c in df.columns}

    def pick(*cands):
        for c in cands:
            if c in cols:
                return cols[c]
        return None

    c_course = pick("course", "course name", "subject")
    c_day    = pick("day", "day of the week", "weekday")
    c_start  = pick("start", "start time", "time from", "from", "time start")
    c_end    = pick("end", "end time", "time to", "to", "time end")

    if not all([c_course, c_day, c_start, c_end]):
        out = df.iloc[:, :4].copy()
        out.columns = ["course", "day", "start", "end"]
        return out

    out = df[[c_course, c_day, c_start, c_end]].copy()
    out.columns = ["course", "day", "start", "end"]
    return out

import pandas as pd

def clean_df(obj) -> pd.DataFrame:
    """
    Ensure we always return a DataFrame with empty rows/cols removed.
    Handles Series or other odd sheet returns gracefully.
    """
    if obj is None:
        return pd.DataFrame()

    # If it's a Series, convert to a 1-column DataFrame
    if isinstance(obj, pd.Series):
        obj = obj.to_frame()

    # If it's not a DataFrame, try converting
    if not isinstance(obj, pd.DataFrame):
        try:
            obj = pd.DataFrame(obj)
        except Exception:
            return pd.DataFrame()

    df = obj.copy()
    df = df.dropna(axis=0, how="all")

    # Only drop empty columns if columns axis exists (DataFrame does)
    df = df.dropna(axis=1, how="all")

    # Normalize column names
    df.columns = [str(c).strip() for c in df.columns]

    return df.reset_index(drop=True)

def pick_best_sheet(sheet_dict: dict) -> tuple[str | None, pd.DataFrame | None]:
    best_name, best_df, best_score = None, None, -1

    for name, obj in sheet_dict.items():
        df2 = clean_df(obj)
        if df2.empty or df2.shape[1] < 4:
            continue

        score = df2.shape[1]
        coltext = " ".join([str(c).lower() for c in df2.columns[:6]])
        if any(k in coltext for k in ["course", "subject", "day", "start", "end", "time"]):
            score += 10

        if score > best_score:
            best_score = score
            best_name = name
            best_df = df2

    return best_name, best_df

def standardize_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Robust standardizer:
    - cleans empty rows/cols
    - if headers don't match, uses first 4 columns
    """
    df = clean_df(df)

    # If after cleaning we still don't have 4 columns, fail with a clearer message
    if df.shape[1] < 4:
        raise ValueError(
            f"Sheet has only {df.shape[1]} non-empty columns after cleaning; "
            "need at least 4: course, day, start, end."
        )

    # Try to detect headers
    cols = {str(c).strip().lower(): c for c in df.columns}

    def pick(*cands):
        for c in cands:
            if c in cols:
                return cols[c]
        return None

    c_course = pick("course", "course name", "subject")
    c_day    = pick("day", "day of the week", "weekday")
    c_start  = pick("start", "start time", "time from", "from", "time start")
    c_end    = pick("end", "end time", "time to", "to", "time end")

    if all([c_course, c_day, c_start, c_end]):
        out = df[[c_course, c_day, c_start, c_end]].copy()
        out.columns = ["course", "day", "start", "end"]
        return out

    # Fallback: assume first 4 columns are the timetable
    out = df.iloc[:, :4].copy()
    out.columns = ["course", "day", "start", "end"]
    return out

def to_minutes(t: time) -> int:
    return t.hour * 60 + t.minute

def floor_minutes(m: int, step: int) -> int:
    return (m // step) * step

def ceil_minutes(m: int, step: int) -> int:
    return ((m + step - 1) // step) * step

def minutes_to_time(m: int) -> time:
    return time(m // 60, m % 60, 0)

def iter_slots(start_min: int, end_min: int, step: int):
    cur = start_min
    while cur < end_min:
        yield cur
        cur += step

def student_color_hex(student_id: str) -> str:
    """
    Deterministic pastel-ish color based on student_id.
    Returns 'RRGGBB'. openpyxl fill wants ARGB => 'FFRRGGBB'.
    """
    h = hashlib.md5(student_id.encode("utf-8")).hexdigest()
    r = int(h[0:2], 16)
    g = int(h[2:4], 16)
    b = int(h[4:6], 16)
    # pastelize
    r = int((r + 255) / 2)
    g = int((g + 255) / 2)
    b = int((b + 255) / 2)
    return f"{r:02X}{g:02X}{b:02X}"


# ==========================================================
# LOAD ALL STUDENT FILES
# ==========================================================
paths = sorted(glob.glob(INPUT_GLOB))
if not paths:
    raise FileNotFoundError(f"No files matched INPUT_GLOB={INPUT_GLOB}")

records = []
students = []

for path in paths:
    skipped = []
    student_id = os.path.splitext(os.path.basename(path))[0]
    students.append(student_id)

    raw = pd.read_excel(path)
    df = standardize_df(raw).dropna(how="all")

    # xl = pd.read_excel(path)  # dict of {sheet_name: df}

    # best_sheet, df_raw = pick_best_sheet(xl)
    # if df_raw is None:
    #     # st.warning(f"Skipping {uploaded_file.name}: couldn't find any sheet with 4 columns (course/day/start/end).")
    #     # continue

    #     # msg = f"Skipping {getattr(uploaded_file, 'name', str(uploaded_file))}: couldn't find any sheet with 4 columns (course/day/start/end)."
    #     # try:
    #     #     import streamlit as st
    #     #     st.warning(msg)
    #     # except Exception:
    #     #     print(msg)
    #     # continue

    #     skipped.append(path)
    #     continue

    
    # df = standardize_df(df_raw)  # your standardize function (updated below)


    df["course"] = df["course"].astype(str).str.strip()
    df["day"] = df["day"].apply(normalize_day)
    df["start"] = df["start"].apply(parse_time)
    df["end"] = df["end"].apply(parse_time)
    df = df.dropna(subset=["course", "day", "start", "end"])
    df = df[df["course"].str.len() > 0].copy()

    for _, r in df.iterrows():
        records.append({
            "Student": student_id,
            "Course": r["course"],
            "Day": r["day"],
            "Start": r["start"],
            "End": r["end"],
        })

if not records:
    raise ValueError("No valid timetable rows found across input files.")

df_all = pd.DataFrame(records)
df_all["Day"] = pd.Categorical(df_all["Day"], categories=DAY_ORDER, ordered=True)
df_all = df_all.sort_values(["Day", "Start", "Student", "Course"]).reset_index(drop=True)

student_list = sorted(set(students))
student_to_color = {s: student_color_hex(s) for s in student_list}


# ==========================================================
# BUILD MATRIX KEYS: (Day, Slot) + student columns
# ==========================================================
min_start = int(df_all["Start"].map(to_minutes).min())
max_end = int(df_all["End"].map(to_minutes).max())

grid_start = floor_minutes(min_start, TIME_STEP_MIN)
grid_end = ceil_minutes(max_end, TIME_STEP_MIN) + TIME_STEP_MIN

days_present = [d for d in DAY_ORDER if d in set(df_all["Day"].astype(str))]
if not days_present:
    days_present = DAY_ORDER[:5]

# Create dict for matrix cell text: (day, slot_min, student) -> list[str]
cell_text = {}
cell_has_class = set()  # (day, slot_min, student)

for _, r in df_all.iterrows():
    d = str(r["Day"])
    if d not in days_present:
        continue

    st = to_minutes(r["Start"])
    en = to_minutes(r["End"])
    st_slot = floor_minutes(st, TIME_STEP_MIN)
    en_slot = ceil_minutes(en, TIME_STEP_MIN)

    s = r["Student"]
    label = f"{r['Course']}\n{r['Start'].strftime('%H:%M')}–{r['End'].strftime('%H:%M')}"

    for slot in iter_slots(st_slot, en_slot, TIME_STEP_MIN):
        key = (d, slot, s)
        cell_text.setdefault(key, []).append(label)
        cell_has_class.add(key)


# ==========================================================
# WRITE EXCEL (Matrix layout)
# Columns: Day | Time | Student1 | Student2 | ...
# ==========================================================
wb = Workbook()
ws = wb.active
ws.title = SHEET_MATRIX
ws_leg = wb.create_sheet(SHEET_LEGEND)

# Styles
thin = Side(style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill("solid", fgColor="F2F2F2")
hdr_font = Font(bold=True)
center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Header
headers = ["Day", "Time"] + student_list
ws.append(headers)

for j, h in enumerate(headers, start=1):
    c = ws.cell(row=1, column=j)
    c.font = hdr_font
    c.fill = hdr_fill
    c.border = border
    c.alignment = center_wrap
    # widths
    if h == "Day":
        ws.column_dimensions[get_column_letter(j)].width = 14
    elif h == "Time":
        ws.column_dimensions[get_column_letter(j)].width = 10
    else:
        ws.column_dimensions[get_column_letter(j)].width = 28

# Rows: iterate Day then slots
row_idx = 2
for day in days_present:
    for slot in iter_slots(grid_start, grid_end, TIME_STEP_MIN):
        ws.cell(row=row_idx, column=1, value=day).border = border
        ws.cell(row=row_idx, column=1).alignment = center_wrap

        t_label = minutes_to_time(slot).strftime("%H:%M")
        ws.cell(row=row_idx, column=2, value=t_label).border = border
        ws.cell(row=row_idx, column=2).alignment = center_wrap

        # student columns
        for k, student in enumerate(student_list, start=3):
            key = (day, slot, student)
            cell = ws.cell(row=row_idx, column=k)

            # default border/alignment
            cell.border = border
            cell.alignment = left_wrap

            if key in cell_has_class:
                # Fill with student's color and write text
                hexrgb = student_to_color[student]
                cell.fill = PatternFill("solid", fgColor="FF" + hexrgb)
                cell.value = "\n".join(cell_text.get(key, []))
                cell.font = Font(bold=True)

        row_idx += 1
    row_idx += 1

# Freeze header & first two cols
ws.freeze_panes = "C2"

# Make it easier to scan: group by day visually (optional row height)
for r in range(2, ws.max_row + 1):
    ws.row_dimensions[r].height = 42

# ==========================================================
# Legend sheet
# ==========================================================
ws_leg.append(["Student", "Color"])
for j in range(1, 3):
    c = ws_leg.cell(row=1, column=j)
    c.font = hdr_font
    c.fill = hdr_fill
    c.border = border
    c.alignment = center_wrap

ws_leg.column_dimensions["A"].width = 45
ws_leg.column_dimensions["B"].width = 18

for i, student in enumerate(student_list, start=2):
    hexrgb = student_to_color[student]
    ws_leg.cell(row=i, column=1, value=student).border = border
    ws_leg.cell(row=i, column=1).alignment = left_wrap

    c = ws_leg.cell(row=i, column=2, value=hexrgb)
    c.fill = PatternFill("solid", fgColor="FF" + hexrgb)
    c.border = border
    c.alignment = center_wrap

ws_leg.freeze_panes = "A2"

# Save
wb.save(OUTPUT_XLSX)
print(f"Saved: {OUTPUT_XLSX}")
